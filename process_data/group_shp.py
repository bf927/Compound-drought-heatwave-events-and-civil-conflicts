from osgeo import ogr
import pandas as pd
import numpy as np
import openpyxl


def saveExcel(array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(array)):
        for jj in range(len(array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = array[ii][jj]  # 写入数据
    workbook.save(save_path)


total_region_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\GeoEPR-2021\\' \
                    'attribute_tabel.xlsx'
conflict_region_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\conflict_data\\' \
                       'select_tabel.xlsx'
conflict_save_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\GeoEPR-2021\\' \
                    'attribute_tabel_with_cf4.xlsx'

total = pd.read_excel(total_region_file)
total_state = np.array(total['statename'])
total_group = np.array(total['group'])
total_year = np.array(total['to'])
total_sqkm = np.array(total['sqkm'])
total = np.array(total)
total[:, -1] = -1

conflict = pd.read_excel(conflict_region_file)
conflict_state = np.array(conflict['statename'])
conflict_group = np.array(conflict['group'])

for i in range(0, conflict_state.shape[0], 1):
    for j in range(0, total_state.shape[0], 1):
        if (total_state[j] == conflict_state[i]) and (total_year[j] >= 2001):
            if total_group[j] == conflict_group[i]:
                total[j, -1] = i
            else:
                continue

conflict_index = total[:, -1]
for i in range(0, conflict_state.shape[0], 1):
    pos = np.where(conflict_index == i)[0]
    if pos.size == 0:
        continue
    else:
        if pos.shape[0] == 1:
            continue
        else:
            sqkm = total_sqkm[pos]
            max_index = np.argmax(sqkm)
            max_pos = pos[max_index]
            total[pos, -1] = -1
            total[max_pos, -1] = i

saveExcel(total, conflict_save_path)