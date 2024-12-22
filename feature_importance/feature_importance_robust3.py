import numpy as np
import pandas as pd
from sklearn.ensemble import RandomForestRegressor
import matplotlib.pyplot as plt
import openpyxl


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


data_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\exp_1989_2020\\' \
            'main_1989_2020_norm_lag.xlsx'
data_path2 = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\exp_1989_2020\\' \
             'main_1989_2020_norm.xlsx'
data = pd.read_excel(data_path)
data2 = pd.read_excel(data_path2)

# 分割数据集为特征和目标变量
X = data[['exclusion', 'his_conf', 'political_group']]

y = data['conflict_frequency']

# 构建随机森林回归模型
rf_model = RandomForestRegressor(n_estimators=500, oob_score=True, verbose=1, n_jobs=-1)
rf_model.fit(X, y)

# 输出各特征的重要性
feature_importance = rf_model.feature_importances_
print(feature_importance)

print("特征重要性：")
for i, feature in enumerate(
        ['exclusion', 'his_conf', 'political_group']):
    print(f"{feature}: {feature_importance[i]}")

weight_array = feature_importance

hs_ssvi = data2['exclusion'] * weight_array[0] + data2['his_conf'] * weight_array[1] + \
          data2['political_group'] * weight_array[2]

array = np.zeros((hs_ssvi.shape[0], 1))
array[:, 0] = hs_ssvi

saveExcel(array,
          'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\exp_1989_2020\\vi.xlsx')

index = np.argsort(feature_importance)
new_feature = np.array([
    'Political exclusion', 'History conflict', 'Political group'])[index]

new_shap = feature_importance[index]

# 绘制每个特征的正负 SHAP 值柱状图
plt.figure(figsize=(16, 10))

# 正值 SHAP 值柱状图
plt.barh(new_feature, new_shap, color='#008080', height=0.35, edgecolor='#2F4F4F', linewidth=2)

font = {'family': 'Arial',
        'color': 'black',
        'weight': 'normal',
        'size': 16,
        }

plt.xlabel("Contribution", fontdict=font)
plt.ylabel("Variable", fontdict=font)
plt.yticks(fontproperties='Arial', size=16)
plt.xticks(fontproperties='Arial', size=16)

ax = plt.gca()
ax.spines['right'].set_color('none')
ax.spines['top'].set_color('none')

plt.savefig('C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\plt\\feature_importance\\'
            '1989_2020.png',
            dpi=600,
            bbox_inches='tight')
# plt.savefig('C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\plt\\feature_importance\\'
#             'all.png',
#             format='png',
#             bbox_inches='tight')
plt.show()
