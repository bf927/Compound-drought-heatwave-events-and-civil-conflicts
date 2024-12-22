import numpy as np
import pandas as pd
import openpyxl


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


table_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\exp_1989_2020\\' \
             'main_1989_2020.xlsx'
save = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\exp_1989_2020\\' \
       'main_1989_2020_norm.xlsx'
table = pd.read_excel(table_path)
data = np.array(table)[:, 6:-2]
print(data.shape)
# print(data)
normalized_data = (data - np.min(data, axis=0)) / (np.max(data, axis=0) - np.min(data, axis=0))
saveExcel(normalized_data, save)