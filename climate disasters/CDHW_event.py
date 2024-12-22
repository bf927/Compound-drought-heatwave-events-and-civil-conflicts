import numpy as np
import glob
from pyrs.algorithm import rs_image
import itertools
import openpyxl
import pandas as pd


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


group_region_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\group_region_file\\' \
                      'raster_filter\\'
Tmax_file = 'D:\\Drought_and_heat_wave_coupling\\data\\Tmax_data\\'
spei_file = 'D:\\Drought_and_heat_wave_coupling\\data\\SPEI_data\\spei\\'
per_25_path = 'D:\\Drought_and_heat_wave_coupling\\data\\percentiles\\25.tiff'
per_75_path = 'D:\\Drought_and_heat_wave_coupling\\data\\percentiles\\75.tiff'
per_90_path = 'D:\\Drought_and_heat_wave_coupling\\data\\percentiles\\90.tiff'
event_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\CDHW_event2\\'
select_tabel_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\conflict_data\\' \
                    'select_tabel.xlsx'

select_tabel = pd.read_excel(select_tabel_path)
state_column = np.array(select_tabel['state index'])
group_column = np.array(select_tabel['group index'])


per_25 = rs_image.Image(per_25_path)
per_25_array = per_25.get_array(True, 1)
per_75 = rs_image.Image(per_75_path)
per_75_array = per_75.get_array(True, 1)
per_90 = rs_image.Image(per_90_path)
per_90_array = per_90.get_array(True, 1)

group_mask_path_list = glob.glob(group_region_file + '*.tif')

for index in range(0, len(group_mask_path_list), 1):
    group_code = int(group_mask_path_list[index][-7:-4])
    print(group_code)
    group_region_data = rs_image.Image(group_mask_path_list[index])
    group_region_array = group_region_data.get_array(True, 1)
    row = np.where(group_region_array == 1)[0]
    column = np.where(group_region_array == 1)[1]

    year_list = glob.glob(spei_file + '*')
    Tmax_year_list = glob.glob(Tmax_file + '*')
    years_list = []
    frequency_list = []
    duration_list = []
    severity_list = []
    for i in range(0, len(year_list), 1):
        year = i + 1989
        print(year)
        days_in_each_month = np.array([31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31])
        days = 365
        if year % 4 == 0:
            days_in_each_month = np.array([31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31])
            days = 366

        month_list = glob.glob(year_list[i] + '\\*.tiff')
        total_month = np.zeros((12, 360, 720), np.float32)
        for j in range(0, len(month_list), 1):
            image = rs_image.Image(month_list[j])
            image_array = image.get_array(True, 1, np.float32)
            total_month[j, :, :] = image_array

        class_value_pixels = total_month[:, row, column]
        class_value_pixels = np.sum(class_value_pixels, axis=1) / class_value_pixels.shape[1]
        yearly_array = np.repeat(class_value_pixels, days_in_each_month)
        yearly_array = np.where(yearly_array > -1, 0, yearly_array)
        yearly_array_dry = np.abs(yearly_array)
        # print(yearly_array)

        Tmax_month_list = glob.glob(Tmax_year_list[i] + '\\*')
        total_day = np.zeros(days)
        day_index = 0
        for j in range(0, len(Tmax_month_list), 1):
            Tmax_day_list = glob.glob(Tmax_month_list[j] + '\\*.tiff')
            for k in range(0, len(Tmax_day_list), 1):
                dataset = rs_image.Image(Tmax_day_list[k])
                array = dataset.get_array(True, 1, np.float32)
                group_array = np.mean(array[row, column])
                group_p90 = np.mean(per_90_array[row, column])
                group_p75 = np.mean(per_75_array[row, column])
                group_p25 = np.mean(per_25_array[row, column])
                if group_array > group_p90:
                    value = (group_array - group_p25) / (group_p75 - group_p25)
                else:
                    value = 0
                total_day[day_index] = value
                day_index += 1

        CDHW_s = total_day * yearly_array_dry
        # print(CDHW_s)

        sequence = np.where(CDHW_s != 0, 1, 0)
        consecutive_ones_events = [list(y) for x, y in itertools.groupby(enumerate(sequence), lambda p: p[1] == 0) if
                                   not x]
        counts = [sum(1 for _, value in event if value == 1) for event in consecutive_ones_events if len(event) >= 3]

        frequency = len(counts)
        duration = sum(counts)
        if frequency != 0:
            severity = np.sum(CDHW_s) / frequency
        else:
            severity = 0

        years_list.append(year)
        frequency_list.append(frequency)
        duration_list.append(duration)
        severity_list.append(severity)

    years_list = np.array([years_list]).T
    frequency_list = np.array([frequency_list]).T
    duration_list = np.array([duration_list]).T
    severity_list = np.array([severity_list]).T

    group_pos = np.where(group_column == group_code)[0]
    state_code = state_column[group_pos]
    group_code_array = np.full(years_list.shape, group_code)
    state_code_array = np.full(years_list.shape, state_code)

    tabel = np.append(years_list, frequency_list, axis=1)
    tabel = np.append(tabel, duration_list, axis=1)
    tabel = np.append(tabel, severity_list, axis=1)
    tabel = np.append(group_code_array, tabel, axis=1)
    tabel = np.append(state_code_array, tabel, axis=1)
    name = np.array([['state', 'group', 'year', 'frequency', 'duration', 'severity']])
    tabel = np.append(name, tabel, axis=0)

    formatted = "{:03}".format(int(group_code))
    path = event_file + formatted + '.xlsx'
    saveExcel(tabel, path)