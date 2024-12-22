import numpy as np
import glob
from pyrs.algorithm import rs_image
import re
import openpyxl


def nan2zero(array):
    nan_pos = np.isnan(array)
    array[nan_pos] = 0
    return array


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


group_region_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\group_region_file\\' \
                    'raster_filter\\'
Tmax_file = 'D:\\Drought_and_heat_wave_coupling\\data\\Tmax_year\\'
spei_file = 'D:\\Drought_and_heat_wave_coupling\\data\\SPEI_data\\year\\'
per_25_path = 'D:\\Drought_and_heat_wave_coupling\\data\\percentiles\\25.tiff'
per_75_path = 'D:\\Drought_and_heat_wave_coupling\\data\\percentiles\\75.tiff'
per_90_path = 'D:\\Drought_and_heat_wave_coupling\\data\\percentiles\\90.tiff'
# save_file_part = 'E:\\CDHW_conflict\\future_climate\\CDHW_feature\\ssp585_part.xlsx'

# save_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\heatwave\\' \
#             'single_heatwave_feature.xlsx'
save_file = 'C:\\Users\\Administrator\\Desktop\\' \
            'heatwave_feature.xlsx'

per_25 = rs_image.Image(per_25_path)
per_25_array = per_25.get_array(True, 1)
per_75 = rs_image.Image(per_75_path)
per_75_array = per_75.get_array(True, 1)
per_90 = rs_image.Image(per_90_path)
per_90_array = per_90.get_array(True, 1)

group_mask_path_list = glob.glob(group_region_file + '*.tif')

group_list = []

length = int(len(group_mask_path_list) * 20)

frequency_array = np.zeros((length, 1))
duration_array = np.zeros((length, 1))
time_array = np.zeros((length, 1))
group_array = np.zeros((length, 1))

for i in range(2001, 2021, 1):
    print(i)
    Tmax_path = Tmax_file + str(i) + '.tif'
    spei_path = spei_file + str(i) + '.tif'

    Tmax_array = rs_image.Image(Tmax_path).get_array(True, 0)
    spei_array = rs_image.Image(spei_path).get_array(True, 0)
    Tmax_array = nan2zero(Tmax_array)
    spei_array = nan2zero(spei_array)
    Tmax_array[Tmax_array == -9999] = 0

    spei_max_value = np.max(spei_array)

    for index in range(0, len(group_mask_path_list), 1):
        group_code = int(group_mask_path_list[index][-7:-4])
        # print(group_code)
        group_region_data = rs_image.Image(group_mask_path_list[index])
        group_region_array = group_region_data.get_array(True, 1)

        tmax_mask = np.prod(Tmax_array, axis=0)

        group_region_array = np.where(tmax_mask == 0, 0, group_region_array)
        group_region_array = np.where(spei_array[0] == spei_max_value, 0, group_region_array)

        row = np.where(group_region_array == 1)[0]
        col = np.where(group_region_array == 1)[1]

        spei_sequence = np.mean(spei_array[:, row, col], axis=1)
        tmax_sequence = np.mean(Tmax_array[:, row, col], axis=1)

        p90 = np.mean(per_90_array[row, col])
        p75 = np.mean(per_75_array[row, col])
        p25 = np.mean(per_25_array[row, col])

        spei_sequence = np.where(spei_sequence > -1, 0, spei_sequence)
        tmax_sequence = np.where(tmax_sequence > p90, (tmax_sequence - p25) / (p75 - p25), 0)
        # cdhw_sequence = tmax_sequence * spei_sequence
        # cdhw_encode = np.where(cdhw_sequence != 0, 1, 0)
        #
        # if cdhw_encode[2] == 0:
        #     cdhw_encode[0] = 0
        #     cdhw_encode[1] = 0
        # if cdhw_encode[-3] == 0:
        #     cdhw_encode[-1] = 0
        #     cdhw_encode[-2] = 0
        #
        # array_str = ''.join(map(str, cdhw_encode))
        #
        # pattern = r'1{3,}(?:0{1}1{3,})*'
        # count_cdhw = len(re.findall(pattern, array_str))
        # array_str = array_str.replace('010', '000').replace('0110', '0000')
        # cdhw_encode = np.frombuffer(array_str.encode('utf-8'), dtype=np.uint8) - ord('0')
        # cdhw_sequence = cdhw_sequence * cdhw_encode

        tmax_encode = np.where(tmax_sequence != 0, 1, 0)
        # tmax_encode = np.where((cdhw_encode == 1) & (tmax_encode == 1), 0, tmax_encode)

        if tmax_encode[2] == 0:
            tmax_encode[0] = 0
            tmax_encode[1] = 0
        if tmax_encode[-3] == 0:
            tmax_encode[-1] = 0
            tmax_encode[-2] = 0

        array_str = ''.join(map(str, tmax_encode))

        pattern = r'1{3,}(?:0{1}1{3,})*'
        count_tmax = len(re.findall(pattern, array_str))
        array_str = array_str.replace('010', '000').replace('0110', '0000')
        tmax_encode = np.frombuffer(array_str.encode('utf-8'), dtype=np.uint8) - ord('0')
        # tmax_sequence = tmax_sequence * tmax_encode

        frequency = count_tmax
        duration = np.sum(tmax_encode)

        pos = int(index * 20 + (i - 2001))
        group_array[pos, :] = group_code
        time_array[pos, :] = i
        frequency_array[pos, :] = frequency
        duration_array[pos, :] = duration

        output_array_part = np.concatenate([group_array, time_array, frequency_array, duration_array],
                                           axis=1)

output_array = np.concatenate([group_array, time_array, frequency_array, duration_array], axis=1)
saveExcel(output_array, save_file)
