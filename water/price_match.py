import pandas as pd
import numpy as np
import openpyxl


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


CROP = 'sunflower'

select_tabel_path = 'D:\\Drought_and_heat_wave_coupling\\data\\water\\HarvestedAreaYield175Crops_Geotiff\\' \
                    'select_tabel.xlsx'
table_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\CDHW\\main_imr.xlsx'
price_path = 'D:\\Drought_and_heat_wave_coupling\\data\\water\\HarvestedAreaYield175Crops_Geotiff\\cash_price_pre\\' \
             + CROP + '\\FAOSTAT.csv'
save = 'D:\\Drought_and_heat_wave_coupling\\data\\water\\HarvestedAreaYield175Crops_Geotiff\\cash_price_pre\\' \
             + CROP + '\\price.xlsx'

state_index = np.array(pd.read_excel(select_tabel_path)['state index'])
state_name = np.array(pd.read_excel(select_tabel_path)['statename'])

output = np.array(pd.read_excel(table_path))[:, 0:4]
output[:, -1] = 0

state_price = np.array(pd.read_csv(price_path)['Area'])
price = np.array(pd.read_csv(price_path)['Value'])

for i in range(0, output.shape[0], 20):
    temp_state_index = output[i, 0]
    print(temp_state_index)
    pos = np.where(state_index == temp_state_index)[0][0]
    print(pos)
    temp_state_name = state_name[pos]
    print(temp_state_name)
    state_pos = np.where(state_price == temp_state_name)[0]
    if state_pos.size == 0:
        continue
    print(state_pos)
    price_state = np.mean(price[state_pos])
    print(price_state)

    output[i:i + 20, -1] = price_state

saveExcel(output, save)
