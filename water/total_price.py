import pandas as pd
import numpy as np
import openpyxl
import glob


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


file = 'D:\\Drought_and_heat_wave_coupling\\data\\water\\HarvestedAreaYield175Crops_Geotiff\\cash_price_pre\\'
file_list = glob.glob(file + '*\\')
# print(file_list)

output = np.zeros((3060, 4))

for i in range(0, len(file_list), 1):
    path = file_list[i] + 'total_price.xlsx'
    data = np.array(pd.read_excel(path))
    output += data
    
saveExcel(output, 'D:\\Drought_and_heat_wave_coupling\\data\\water\\HarvestedAreaYield175Crops_Geotiff\\'
                  'cash_price_pre\\total.xlsx')