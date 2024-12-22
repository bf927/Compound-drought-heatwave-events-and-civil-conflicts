from pyrs.algorithm import rs_image
import numpy as np
import cv2
import pandas as pd
import openpyxl

# real


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


def nan2zero(array):
    nan_pos = np.isnan(array)
    array[nan_pos] = 0
    return array


def negative2zero(array):
    array = np.where(array < 0, 0, array)
    return array


etp_file = 'D:\\Drought_and_heat_wave_coupling\\data\\water\\etp\\rainfall\\GW\\'
eta_file = 'D:\\Drought_and_heat_wave_coupling\\data\\water\\eta\\rainfall\\GW\\'
rainfall_area_file = 'D:\\Drought_and_heat_wave_coupling\\data\\water\\crop_area\\rainfall\\'
rainfall_max_yield_file = 'D:\\Drought_and_heat_wave_coupling\\data\\water\\HarvestedAreaYield175Crops_Geotiff\\' \
                          'yield_26\\'
ky_path = 'D:\\Drought_and_heat_wave_coupling\\data\\water\\HarvestedAreaYield175Crops_Geotiff\\ky.xlsx'
save_file = 'D:\\Drought_and_heat_wave_coupling\\data\\water\\HarvestedAreaYield175Crops_Geotiff\\' \
            'yield_error3\\'

ky_array = np.array(pd.read_excel(ky_path)['ky']).flatten()

for i in range(1, 27, 1):
    print(i)
    formatted = "{:02}".format(int(i))
    save_path = save_file + formatted + '.xlsx'

    error_array = np.zeros((20, 3))
    index = 0
    ky = ky_array[i - 1]
    print(ky)
    for year in range(2001, 2021, 1):
        print(year)

        np.seterr(all="ignore")
        formatted = "{:02}".format(int(i))

        etp_path = etp_file + formatted + '\\' + str(year) + '.tif'
        eta_path = eta_file + formatted + '\\' + str(year) + '.tif'
        rainfall_area_path = rainfall_area_file + '\\annual_area_harvested_rfc_crop' + formatted + '_ha_30mn1.tif'
        rainfall_max_yield_path = rainfall_max_yield_file + formatted + '.tif'

        etp = rs_image.Image(etp_path).get_array(True, 1)
        eta = rs_image.Image(eta_path).get_array(True, 1)
        eta_15p = 1.15 * eta
        eta_15n = 0.85 * eta

        rainfall_area = nan2zero(rs_image.Image(rainfall_area_path).get_array(True, 1))
        rainfall_max_yield = nan2zero(rs_image.Image(rainfall_max_yield_path).get_array(True, 1))

        rainfall_max_yield_resample = cv2.resize(rainfall_max_yield, (720, 360), interpolation=cv2.INTER_NEAREST)

        ks = eta / etp
        new_ks = nan2zero(ks)

        ks_15p = eta_15p / etp
        new_ks_15p = nan2zero(ks_15p)

        ks_15n = eta_15n / etp
        new_ks_15n = nan2zero(ks_15n)

        # print(rainfall_max_yield_resample.shape)
        # print(ky)
        # print(new_ks)

        yield_act = rainfall_max_yield_resample * (1 + ky * (new_ks - 1))
        yield_act_15p = rainfall_max_yield_resample * (1 + ky * (new_ks_15p - 1))
        yield_act_15n = rainfall_max_yield_resample * (1 + ky * (new_ks_15n - 1))

        mask = np.where(rainfall_area == 0, 0, 1)

        yield_act_filter = np.where(yield_act * mask < 0, 0, yield_act * mask)
        yield_act_15p_filter = np.where(yield_act_15p * mask < 0, 0, yield_act_15p * mask)
        yield_act_15n_filter = np.where(yield_act_15n * mask < 0, 0, yield_act_15n * mask)

        act = np.sum(yield_act_filter) / np.sum(yield_act_filter != 0)
        act_15p = np.sum(yield_act_15p_filter) / np.sum(yield_act_15p_filter != 0)
        act_15n = np.sum(yield_act_15n_filter) / np.sum(yield_act_15n_filter != 0)

        error_15p = np.abs(act - act_15p) / act
        error_15n = np.abs(act - act_15n) / act

        error_array[index, 0] = act
        error_array[index, 1] = error_15p
        error_array[index, 2] = error_15n

        index += 1

    act_array = error_array[:, 0]
    indices = np.argsort(act_array)
    error_array = error_array[indices, :]
    if ky == 1:
        error_array[:, 1] = np.mean(error_array[:, 1])
        error_array[:, 2] = np.mean(error_array[:, 2])

    name = np.array([['yield', '15p', '15n']])
    error_array = np.append(name, error_array, axis=0)
    saveExcel(error_array, save_path)
