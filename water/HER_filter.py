import numpy as np
import pandas as pd
import openpyxl


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\death_HW\\main_hw_death3.xlsx'
data = np.array(pd.read_excel(path)['energy_cap'])

her = np.where(data < 1800, 1, 0).reshape((-1, 1))
saveExcel(her, 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\death_HW\\her.xlsx')