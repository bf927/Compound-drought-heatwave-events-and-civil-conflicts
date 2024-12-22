import glob
from pyrs.algorithm import rs_image
import numpy as np
import pandas as pd
import openpyxl


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


def nan2zero(array):
    nan_pos = np.isnan(array)
    array[nan_pos] = 0
    return array


def watneeds(pet_file, pre_file, infiltration_path, awc_path, z_p_data_path, kc_path, kc_BIS_path, crop_area_path,
             output_file, output_file2, output_file3):
    # 土壤最大下渗率
    f_dataset = rs_image.Image(infiltration_path)
    f = f_dataset.get_array(True, 1)

    # 土壤最大含水量
    awc_dataset = rs_image.Image(awc_path)
    awc = awc_dataset.get_array(True, 1)
    awc = nan2zero(awc)

    # z和p
    table = pd.read_excel(z_p_data_path)
    if TYPE == 'rainfall':
        z = np.array(table['rainfall'])[CROP_ID - 1]
    else:
        z = np.array(table['irritated'])[CROP_ID - 1]
    p = np.array(table['p'])[CROP_ID - 1]

    # area
    crop_area_dataset = rs_image.Image(crop_area_path)
    crop_area_array = crop_area_dataset.get_array(True, 1) * 0.01  # 公顷转换为平方千米
    crop_area_array = nan2zero(crop_area_array)
    # print(np.sum(crop_area_array))

    # readily available water
    raw = p * awc * z
    taw = awc * z

    # 1999-2020年天数
    days_in_each_year = [365, 366, 365, 365, 365, 366, 365, 365, 365, 366, 365, 365, 365, 366, 365, 365, 365, 366, 365,
                         365, 365, 366]

    total_days = sum(days_in_each_year)

    # 初始化
    s = np.zeros((total_days, 360, 720), dtype=np.float16)

    n = 0  # 20年的累计天数
    flag = 0
    for i in range(0, len(days_in_each_year), 1):  # 遍历每一年
        year = i + 1999
        pet_year_file = pet_file + str(year)
        pet_path_list = glob.glob(pet_year_file + '\\*.tif')

        pet_month = np.zeros((12, 360, 720), np.float32)  # 构造特定年的月度pet数组
        for j in range(0, len(pet_path_list), 1):
            image = rs_image.Image(pet_path_list[j])
            image_array = image.get_array(True, 1, np.float32)
            pet_month[j, :, :] = image_array

        if year % 4 != 0:  # 判断是否为闰年
            days_in_month = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
            kc_dataset = rs_image.Image(kc_path)
            kc_daily = kc_dataset.get_array(True, 0)
        else:
            days_in_month = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
            kc_BIS_dataset = rs_image.Image(kc_BIS_path)
            kc_daily = kc_BIS_dataset.get_array(True, 0)

        # 初始化一个新的数组，用于存放每天的数据
        pet_daily = np.ones((sum(days_in_month), 360, 720), dtype=pet_month.dtype)

        # 记录上一个月份结束时的累积天数
        sum_days = 0
        for month, days in enumerate(days_in_month):
            pet_daily[sum_days: sum_days + days] = np.repeat(pet_month[month][np.newaxis, ...],
                                                             days, axis=0)  # 将当前月份的数据复制到daily_data数组中
            sum_days += days  # 更新累积天数

        pre_year_file = pre_file + str(year)
        pre_path_list = glob.glob(pre_year_file + '\\*.tif')

        annual_GW = np.zeros((days_in_each_year[i], 360, 720))
        annual_GW_stress = np.zeros((days_in_each_year[i], 360, 720))
        annual_BW = np.zeros((days_in_each_year[i], 360, 720))

        for t in range(0, days_in_each_year[i], 1):
            # print(t)
            np.seterr(all="ignore")

            if n == 0:
                flag = 1
                s[n, :, :] = 0.5 * taw
                n = 1
                t = 1

            pre = rs_image.Image(pre_path_list[t]).get_array(True, 1)  # 每日降水量
            pre_eff = pre * 0.95  # 每日有效降水量(95%)
            et0 = pet_daily[t, :, :]  # 每日参考蒸散发
            kc = kc_daily[t, :, :]  # 每日作物系数

            et = et0 * kc  # 每日潜在蒸散发

            ks = np.where(s[n - 1, :, :] >= raw, 1, s[n - 1, :, :] / raw)  # 计算胁迫系数
            gw = ks * et  # 绿水

            dp = np.where(s[n - 1, :, :] >= raw, f * (s[n - 1, :, :] - raw) / (taw - raw), 0)
            s[n, :, :] = s[n - 1, :, :] + pre_eff - gw - dp

            gw_ratio = gw / (gw + dp)
            dp_ratio = dp / (gw + dp)
            gw = np.where(s[n, :, :] < 0, gw_ratio * (s[n - 1, :, :] + pre_eff), gw)
            dp = np.where(s[n, :, :] < 0, dp_ratio * (s[n - 1, :, :] + pre_eff), dp)

            s[n, :, :] = s[n - 1, :, :] + pre_eff - gw - dp

            s[n, :, :] = np.where(s[n, :, :] > taw, taw, s[n, :, :])
            annual_GW[t, :, :] = et
            annual_GW_stress[t, :, :] = gw
            if TYPE == 'irritated':
                bw = et - gw
                annual_BW[t, :, :] = bw

            if flag == 1:
                n = 0
                flag = 0
            n += 1

        annual_total_GW = np.sum(annual_GW, axis=0) * crop_area_array / 1000000
        annual_total_GW = nan2zero(annual_total_GW)
        annual_total_GW_stress = np.sum(annual_GW_stress, axis=0) * crop_area_array / 1000000
        annual_total_GW_stress = nan2zero(annual_total_GW_stress)

        print(str(year) + ' GW ' + str(np.sum(annual_total_GW_stress)))
        print(str(year) + ' GW NO Stress ' + str(np.sum(annual_total_GW)))

        GW_stress_path = output_file + str(year) + '.tif'
        GW_no_stress_path = output_file3 + str(year) + '.tif'
        awc_dataset.save(GW_stress_path, annual_total_GW_stress)
        awc_dataset.save(GW_no_stress_path, annual_total_GW)

        if TYPE == 'irritated':
            BW_path = output_file2 + str(year) + '.tif'
            annual_total_BW = np.sum(annual_BW, axis=0) * crop_area_array / 1000000
            nan2zero(annual_total_BW)

            print(str(year) + ' BW ' + str(np.sum(annual_total_BW)))

            awc_dataset.save(BW_path, annual_total_BW)


if __name__ == '__main__':
    TYPE = 'rainfall'
    TYPE2 = 'rfc'
    WATER_TP = 'GW'
    WATER_TP2 = 'BW'

    # TYPE = 'irritated'
    # TYPE2 = 'irc'
    # WATER_TP = 'GW'
    # WATER_TP2 = 'BW'

    crop_gw_list = []

    for CROP_ID in range(1, 27, 1):
        formatted = "{:02}".format(int(CROP_ID))
        print(formatted)

        pet_file = 'D:\\Drought_and_heat_wave_coupling\\data\\water\\pet_image\\'
        pre_file = 'D:\\Drought_and_heat_wave_coupling\\data\\water\\pre_image\\'
        infiltration_path = 'D:\\Drought_and_heat_wave_coupling\\data\\water\\infiltration_rate\\infiltration.tif'
        awc_path = 'D:\\Drought_and_heat_wave_coupling\\data\\water\\awc\\awc.tif'
        # awc_path = 'D:\\Drought_and_heat_wave_coupling\\data\\water\\awc\\' \
        #            'sol_available.water.capacity_usda.mm_m_250m_0..200cm.tif'
        z_p_data_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\water\\z_p_2.xlsx'
        kc_path = 'D:\\Drought_and_heat_wave_coupling\\data\\water\\kc2\\' + TYPE + '\\no_BIS\\' + formatted + '.tif'
        kc_BIS_path = 'D:\\Drought_and_heat_wave_coupling\\data\\water\\kc2\\' + TYPE + '\\BIS\\' + formatted + '.tif'
        crop_area_path = 'D:\\Drought_and_heat_wave_coupling\\data\\water\\crop_area\\' + TYPE + \
                         '\\annual_area_harvested_' + TYPE2 + '_crop' + formatted + '_ha_30mn1.tif'
        output_file = 'D:\\Drought_and_heat_wave_coupling\\data\\water\\eta\\' + TYPE + '\\' + WATER_TP + '\\' + \
                      "{:02}".format(int(CROP_ID)) + '\\'
        output_file2 = 'D:\\Drought_and_heat_wave_coupling\\data\\water\\eta\\' + TYPE + '\\' + WATER_TP2 + '\\' + \
                       "{:02}".format(int(CROP_ID)) + '\\'
        output_file3 = 'D:\\Drought_and_heat_wave_coupling\\data\\water\\etp\\' + TYPE + '\\' + WATER_TP + '\\' + \
                       "{:02}".format(int(CROP_ID)) + '\\'

        watneeds(pet_file, pre_file, infiltration_path, awc_path, z_p_data_path, kc_path, kc_BIS_path, crop_area_path,
                 output_file, output_file2, output_file3)