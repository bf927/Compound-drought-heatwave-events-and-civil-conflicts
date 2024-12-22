import pandas as pd
import numpy as np
import openpyxl


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\CDHW_conflict_factor.xlsx'
data = np.array(pd.read_excel(path))
array = data[:, 6:-4]
# print(array)

array_list = []
for i in range(0, array.shape[0], 20):
    group_array = array[i:i+20, :]
    group_array_2001_2019 = group_array[0:19, :]
    group_array_2002_2020 = group_array[1:20, :]
    group_change_array = group_array_2002_2020 - group_array_2001_2019
    array_list.append(group_change_array)

new_array = array_list[0]
for i in range(1, len(array_list), 1):
    new_array = np.append(new_array, array_list[i], axis=0)

saveExcel(new_array,
          'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\CDHW_conflict_change.xlsx')