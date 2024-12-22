import pandas as pd
import numpy as np
import glob
import openpyxl


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\CDHW_event2\\'
path_list = glob.glob(file + '*.xlsx')
print(path_list)
array_list = []

group_region_file = 'D:\\Drought_and_heat_wave_coupling\\data\\mask\\gw\\raster_filter\\'
group_mask_path_list = glob.glob(group_region_file + '*.tif')
group_code_list = []

for i in range(0, len(group_mask_path_list), 1):
    group_code = int(group_mask_path_list[i][-7:-4])
    group_code_list.append(group_code)

group_code_array = np.array(group_code_list)

for i in range(0, len(path_list), 1):
    index = int(path_list[i][-8:-5])
    print(path_list[i][-8:-5])
    if np.where(group_code_array == index)[0].size == 0:
        continue
    data = pd.read_excel(path_list[i])
    data = np.array(data)
    array_list.append(data)

new_array = array_list[0]
for i in range(1, len(array_list), 1):
    new_array = np.append(new_array, array_list[i], axis=0)

saveExcel(new_array,
          'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\exp_1989_2020\\CDHW.xlsx')