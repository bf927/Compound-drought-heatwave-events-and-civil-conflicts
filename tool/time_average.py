import pandas as pd
import numpy as np
import openpyxl


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\reg_feature\\main.xlsx'
data = np.array(pd.read_excel(path))
# print(data)

YEAR = 20

average_data = np.zeros((data.shape[0] // YEAR, data.shape[1]))
print(average_data.shape)

num = 0
for i in range(0, data.shape[0], YEAR):
    array = np.mean(data[i:i + YEAR, :], axis=0)
    average_data[num, :] = array
    num += 1

print(average_data)
saveExcel(average_data,
          'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\reg_feature\\main_time_avg.xlsx')