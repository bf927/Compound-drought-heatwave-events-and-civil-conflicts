import pandas as pd
import numpy as np
import openpyxl


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\reg_feature\\main.xlsx'
data = np.array(pd.read_excel(path))
# print(data)

average_data = np.zeros((20, data.shape[1]))

for i in range(0, data.shape[0], 20):
    array = data[i:i + 20, :]
    average_data += array

# print(average_data)
average_data = average_data / (data.shape[0] // 20)
saveExcel(average_data,
          'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\reg_feature\\main_avg.xlsx')