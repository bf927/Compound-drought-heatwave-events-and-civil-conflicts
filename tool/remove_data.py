import numpy as np
import pandas as pd
import openpyxl


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


table_path = 'D:\\Drought_and_heat_wave_coupling\\data\\cumNDVI\\ad.xlsx'
table = pd.read_excel(table_path)
table = np.array(table)
index = np.unique(table[:, 0])

table2_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\exclusion\\total.xlsx'
table2 = pd.read_excel(table2_path)
table2 = np.array(table2)
index2 = table2[:, 0]

pos = np.isin(index2, index, invert=True)
new_array = np.delete(table2, pos, axis=0)
saveExcel(new_array,
          'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\exclusion\\exclusion.xlsx')