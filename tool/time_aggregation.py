import pandas as pd
import numpy as np
import openpyxl
import glob

k = 6


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


save_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\var_conf\\'
tabel_list = glob.glob(save_file + '*')
new_save_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\var_conf_time2\\'

for i in range(0, len(tabel_list), 1):
    data = pd.read_excel(tabel_list[i])
    data = np.array(data)
    stata_name = data[0, 0]
    group_name = data[0, 1]

    feature_array = np.array(data[:, 3:6], dtype=np.float64)
    control_array = np.array(data[:, 6:10], dtype=np.float64)
    conflict_array = np.array(data[:, 10], dtype=np.float64)

    feature_list = []
    control_list = []
    conflict_list = []
    for row in range(0, data.shape[0], k):
        sub_feature = feature_array[row:row + k, :]
        feature_list.append(sub_feature)

        sub_control = control_array[row:row + k, :]
        control_list.append(sub_control)

        sub_conflict = conflict_array[row:row + k]
        conflict_list.append(sub_conflict)

    # 处理最后一个子数组，如果不足 k 行则加入
    if data.shape[0] % k != 0:
        feature_list.append(feature_array[-(data.shape[0] % k):, :])
        control_list.append(control_array[-(data.shape[0] % k):, :])
        conflict_list.append(conflict_array[-(data.shape[0] % k):])

    new_array = np.zeros((len(feature_list), 9))
    for j in range(0, len(feature_list), 1):
        new_array[j, 0] = j
        new_array[j, 1:4] = np.sum(feature_list[j], axis=0)
        new_array[j, 4:8] = np.mean(control_list[j], axis=0)
        if np.sum(conflict_list[j]) != 0:
            new_array[j, 8] = 1
        else:
            new_array[j, 8] = 0

    stata_array = np.full((len(feature_list), 1), stata_name)
    group_array = np.full((len(feature_list), 1), group_name)

    f = new_array[:, 1]
    pos = np.where(f == 0)[0]
    if pos.size != 0:
        continue

    name = np.array([['stata', 'group', 'year', 'frequency', 'duration', 'severity', 'ad', 'GDP', 'population',
                      'IMR', 'conflict']])
    new_array = np.append(group_array, new_array, axis=1)
    new_array = np.append(stata_array, new_array, axis=1)
    new_array = np.append(name, new_array, axis=0)
    save_name = tabel_list[i][-8:]
    save_path = new_save_file + save_name
    saveExcel(new_array, save_path)