import numpy as np
import pandas as pd
import openpyxl


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


path1 = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\extreme climate\\p3_p4_hsvi.xlsx'
path2 = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\extreme climate\\' \
        'p3_p4_duration.xlsx'

data1 = pd.read_excel(path1)
data2 = pd.read_excel(path2)

data1_group = np.unique(data1['group'])
data2_group = np.unique(data2['group'])

group = np.array(data2['group'])

intersection = np.array(np.intersect1d(data1_group, data2_group))
if intersection.size == 0:
    print('NAN')
else:
    print(intersection)
    pos = np.where(np.in1d(group, intersection))[0]
    intersection_data = np.array(data2)[pos, :]
    print(intersection_data.shape)
    saveExcel(intersection_data, 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\'
                                 'extreme climate\\intersect_hsvi_duration.xlsx')