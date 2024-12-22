from pyrs.algorithm import supervision_classification as sc
from pyrs.algorithm import rs_image
from osgeo import gdal
import numpy as np
import glob
from sklearn.linear_model import LinearRegression
import openpyxl
import pandas as pd


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


group_region_file = 'D:\\Drought_and_heat_wave_coupling\\data\\mask\\country\\raster_filter\\'
group_mask_path_list = glob.glob(group_region_file + '*.tif')

image_file = 'D:\\Drought_and_heat_wave_coupling\\data\\landcover\\'
ndvi_file = 'D:\\Drought_and_heat_wave_coupling\\data\\cumNDVI\\'

landcover_list = glob.glob(image_file + '*.tif')
ndvi_list = glob.glob(ndvi_file + '*.tif')

array_list = []
for i in range(0, len(group_mask_path_list), 1):
    group_code = int(group_mask_path_list[i][-7:-4])
    group_region_data = rs_image.Image(group_mask_path_list[i])
    mask_array = group_region_data.get_array(True, 1)
    print(group_code)

    rainfall_list = []
    irrigate_list = []
    for j in range(0, len(landcover_list), 1):
        image = rs_image.Image(landcover_list[j])
        image_array = image.get_array(True, 1)
        ndvi = rs_image.Image(ndvi_list[j])
        ndvi_array = ndvi.get_array(True, 1) * 0.0001
        nan_pos = np.isnan(ndvi_array)
        ndvi_array[nan_pos] = 0

        image_array = np.where(image_array == 11, 10, image_array)
        image_array = np.where(image_array == 12, 10, image_array)

        rainfall_crop_index = np.where((mask_array == 1) & (image_array == 10))
        rainfall_crop_ndvi = np.sum(ndvi_array[rainfall_crop_index])
        rainfall_list.append(rainfall_crop_ndvi)

        irrigate_crop_index = np.where((mask_array == 1) & (image_array == 20))
        irrigate_crop_ndvi = np.sum(ndvi_array[irrigate_crop_index])
        irrigate_list.append(irrigate_crop_ndvi)

    year = np.arange(2001, 2021).reshape((-1, 1))
    rainfall_value = np.array(rainfall_list).reshape((-1, 1))
    irrigate_value = np.array(irrigate_list).reshape((-1, 1))

    index = np.full((20, 1), group_code)
    array = np.append(index, rainfall_value, axis=1)
    array = np.append(array, irrigate_value, axis=1)
    array_list.append(array)

array = array_list[0]
for i in range(1, len(array_list), 1):
    array = np.append(array, array_list[i], axis=0)

saveExcel(array, 'D:\\Drought_and_heat_wave_coupling\\data\\cumNDVI\\cumNDVI_civilian.xlsx')
