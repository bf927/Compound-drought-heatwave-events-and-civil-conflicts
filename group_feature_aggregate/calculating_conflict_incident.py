import pandas as pd
import numpy as np
from pyrs.algorithm import rs_image
import openpyxl
import glob


def saveExcel(Array, savePath):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(savePath)


conflict_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\conflict_data\\' \
                'country_region_conflict\\'
save_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\var_conf\\'
group_region_file = 'D:\\Drought_and_heat_wave_coupling\\data\\mask\\country\\raster_filter\\'


group_mask_path_list = glob.glob(group_region_file + '*.tif')


for i in range(0, len(group_mask_path_list), 1):
    group_code = int(group_mask_path_list[i][-7:-4])
    print(group_code)

    conflict_path = conflict_file + str(int(group_code)) + '.xlsx'
    conflict = pd.read_excel(conflict_path)
    conflict_array = np.array(conflict)
    years = conflict_array[:, 2]
    side = conflict_array[:, 15]
    conflict_list = []
    for j in range(2000, 2020, 1):
        if np.sum(years == j) > 0:
            conflict_list.append(1)
        else:
            conflict_list.append(0)

    conflict_list = np.array([conflict_list]).T
    index = np.full((20, 1), int(group_code))
    event_array = np.append(index, conflict_list, axis=1)

    name = np.array([['group', 'conflict']])
    event_array = np.append(name, event_array, axis=0)

    formatted = "{:03}".format(int(group_code))
    save_path = save_file + formatted + '.xlsx'
    saveExcel(event_array, save_path)