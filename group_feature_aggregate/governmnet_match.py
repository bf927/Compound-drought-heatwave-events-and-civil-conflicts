import pandas as pd
import numpy as np
import openpyxl


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


select_tabel_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\government efficiency\\' \
                    'select_tabel.xlsx'
table_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\CDHW\\main_imr.xlsx'
government_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\government efficiency\\' \
                  'Voice and Accountability.xlsx'
save = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\government efficiency\\' \
       'Voice and Accountability group.xlsx'

state_index = np.array(pd.read_excel(select_tabel_path)['state index'])
state_name = np.array(pd.read_excel(select_tabel_path)['code'])

output = np.array(pd.read_excel(table_path))[:, 0:4]
output[:, -1] = 0

index_list = np.array(np.arange(2, 122, 6), dtype=np.int)

state_government = np.array(pd.read_excel(government_path)['Code'])
government_efficiency = np.array(pd.read_excel(government_path))[:, index_list]

for i in range(0, output.shape[0], 20):
    temp_state_index = output[i, 0]
    print(temp_state_index)
    pos = np.where(state_index == temp_state_index)[0][0]
    print(pos)
    temp_state_name = state_name[pos]
    print(temp_state_name)
    state_pos = np.where(state_government == temp_state_name)[0]
    if state_pos.size == 0:
        continue
    print(state_pos)
    death_rate_state = government_efficiency[state_pos, :]
    print(death_rate_state)

    output[i:i + 20, -1] = death_rate_state

saveExcel(output, save)