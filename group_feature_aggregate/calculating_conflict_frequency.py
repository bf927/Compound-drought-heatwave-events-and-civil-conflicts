import pandas as pd
import numpy as np
import openpyxl
import glob


def saveExcel(Array, savePath):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(savePath)


conflict_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\conflict_data\\' \
                'group_region_conflict\\'
save_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\exp_1989_2020\\' \
            'group_conflict_frequency\\'
group_region_file = 'D:\\Drought_and_heat_wave_coupling\\data\\mask\\gw\\raster_filter\\'

group_mask_path_list = glob.glob(group_region_file + '*.tif')

array_list = []
for i in range(0, len(group_mask_path_list), 1):
    group_code = int(group_mask_path_list[i][-7:-4])
    print(group_code)

    conflict_path = conflict_file + str(int(group_code)) + '.xlsx'
    conflict = pd.read_excel(conflict_path)
    conflict_array = np.array(conflict)
    years = conflict_array[:, 2]

    conflict_list = []
    for j in range(1989, 2021, 1):
        if np.sum(years == j) > 0:
            conflict_list.append(np.sum(years == j))
        else:
            conflict_list.append(0)

    conflict_list = np.array([conflict_list]).T
    index = np.full((2020 - 1989 + 1, 1), int(group_code))
    year_array = np.arange(1989, 2021, 1).reshape((-1, 1))
    event_array = np.append(year_array, conflict_list, axis=1)
    event_array = np.append(index, event_array, axis=1)

    array_list.append(event_array)

array = array_list[0]
for i in range(1, len(array_list), 1):
    array = np.append(array, array_list[i], axis=0)

saveExcel(array, 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\exp_1989_2020\\'
                 'group_conflict_frequency.xlsx')
