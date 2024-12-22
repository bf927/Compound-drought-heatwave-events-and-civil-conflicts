from pyrs.algorithm import supervision_classification as sc
from pyrs.algorithm import rs_image
from osgeo import gdal
import numpy as np
import glob
from sklearn.linear_model import LinearRegression
import openpyxl
import pandas as pd


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


mask_path = 'D:\\Drought_and_heat_wave_coupling\\data\\crop_production\\group_mask.tif'
mask_image = rs_image.Image(mask_path)
mask_array = mask_image.get_array(True, 1)
group_value = np.unique(mask_array)
group_value = group_value[1:]

TWS_file = 'D:\\Drought_and_heat_wave_coupling\\data\\TWS\\'

TWS_list = glob.glob(TWS_file + '*.tif')

array_list = []
for i in range(0, group_value.shape[0], 1):
    print(group_value[i])
    mean_TWS_list = []
    for j in range(0, len(TWS_list), 1):
        image = rs_image.Image(TWS_list[j])
        image_array = image.get_array(True, 1)
        TWS = rs_image.Image(TWS_list[j])
        TWS_array = TWS.get_array(True, 1)

        index = np.where(mask_array == group_value[i])
        mean_TWS = np.mean(TWS_array[index])
        mean_TWS_list.append(mean_TWS)

    year = np.arange(2001, 2021).reshape((-1, 1))
    TWS_value = np.array(mean_TWS_list).reshape((-1, 1))

    index = np.full((20, 1), group_value[i])
    array = np.append(index, TWS_value, axis=1)
    array_list.append(array)

array = array_list[0]
for i in range(1, len(array_list), 1):
    array = np.append(array, array_list[i], axis=0)

saveExcel(array, 'D:\\Drought_and_heat_wave_coupling\\data\\TWS\\TWS.xlsx')
