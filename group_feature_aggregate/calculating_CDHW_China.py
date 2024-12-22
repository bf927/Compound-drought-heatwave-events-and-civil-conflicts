from pyrs.algorithm import rs_image
import numpy as np
import glob
import openpyxl


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


group_region_path = 'D:\\Drought_and_heat_wave_coupling\\data\\mask\\china\\01.tif'
image_file = 'E:\\CDHW_conflict\\future_climate\\CDHW_image\\ssp585\\duration\\'

image_list = glob.glob(image_file + '*.tif')


group_region_data = rs_image.Image(group_region_path)
mask_array = group_region_data.get_array(True, 1)

value_list = []
for j in range(0, len(image_list), 1):
    image = rs_image.Image(image_list[j])
    image_array = image.get_array(True, 1)
    nan_index = np.isnan(image_array)
    image_array[nan_index] = 0

    group_area_index = np.where(mask_array == 1)
    group_light = np.mean(image_array[group_area_index])
    value_list.append(group_light)

year = np.arange(2021, 2101).reshape((-1, 1))
area_value = np.array(value_list).reshape((-1, 1))

index = np.full((80, 1), 0)
array = np.append(index, area_value, axis=1)

saveExcel(array, 'E:\\CDHW_conflict\\future_climate\\CDHW_image\\ssp585\\duration\\china.xlsx')