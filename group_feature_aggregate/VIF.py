import pandas as pd
from statsmodels.stats.outliers_influence import variance_inflation_factor
import numpy as np
import openpyxl


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


def calculate_vif(data_frame):
    # 创建一个DataFrame，其中包含所有独立变量的名称和它们的VIF值
    vif_data = pd.DataFrame()
    vif_data["Variable"] = data_frame.columns
    vif_data["VIF"] = [variance_inflation_factor(data_frame.values, i) for i in range(data_frame.shape[1])]

    return vif_data


# 示例
data_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\VIF\\VIF_f_d.xlsx'

df = pd.read_excel(data_path)

vif_result = np.array(calculate_vif(df))
saveExcel(vif_result, 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\VIF\\'
                      'VIF_result_f_d.xlsx')
print(vif_result)
