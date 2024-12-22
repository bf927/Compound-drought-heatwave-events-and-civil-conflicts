from pyrs.algorithm import supervision_classification as sc
from pyrs.algorithm import rs_image
from osgeo import gdal
import numpy as np
import glob
from sklearn.linear_model import LinearRegression
import openpyxl


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


# path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\GeoEPR-2021\\Export_Output.shp'
# ex_path = 'D:\\Drought_and_heat_wave_coupling\\data\\lighting_data\\F101992.v4\\' \
#           'F101992.v4b_web.stable_lights.avg_vis.tif'
# ex_image = rs_image.Image(ex_path)
# array = sc.shapefile_to_raster(path, ex_image.height, ex_image.width, ex_image.proj, ex_image.geotrans,
#                                'conflict', gdal.GDT_Float32)
# array = np.where(array == 0, -1, array)
# array = np.where(array == 170, 0, array)
# ex_image.save('D:\\Drought_and_heat_wave_coupling\\data\\lighting_data\\mask.tif', array)


mask_path = 'D:\\Drought_and_heat_wave_coupling\\data\\lighting_data\\resample_mask.tif'
mask_image = rs_image.Image(mask_path)
mask_array = mask_image.get_array(True, 1)
group_value = np.unique(mask_array)
group_value = group_value[1:]
# print(group_value)
image_file = 'D:\\Drought_and_heat_wave_coupling\\data\\lighting_data\\resample\\'

landcover_list = glob.glob(image_file + '*.tif')
# print(landcover_list)

array_list = []
for i in range(0, group_value.shape[0], 1):
    ad_list = []
    for j in range(0, len(landcover_list), 1):
        image = rs_image.Image(landcover_list[j])
        image_array = image.get_array(True, 1)
        group_area_index = np.where(mask_array == group_value[i])
        group_light = np.log(np.sum(image_array[group_area_index]) + 0.0001)
        ad_list.append(group_light)

    year = np.arange(1992, 2014).reshape((-1, 1))
    ad_value = np.array(ad_list)
    regression_model = LinearRegression()
    regression_model.fit(year, ad_value)

    year_pass = np.arange(1989, 1992).reshape((-1, 1))
    year_future = np.arange(2014, 2023).reshape((-1, 1))
    ad_pass = regression_model.predict(year_pass)
    ad_future = regression_model.predict(year_future)
    ad_value = np.append(ad_pass, ad_value)
    ad_value = np.append(ad_value, ad_future).reshape((-1, 1))
    index = np.full((34, 1), group_value[i])
    array = np.append(index, ad_value, axis=1)
    # print(array.shape)
    array_list.append(array)

array = array_list[0]
for i in range(1, len(array_list), 1):
    array = np.append(array, array_list[i], axis=0)

saveExcel(array, 'D:\\Drought_and_heat_wave_coupling\\data\\lighting_data\\light.xlsx')