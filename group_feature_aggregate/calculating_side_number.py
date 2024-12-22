import pandas as pd
import numpy as np
from pyrs.algorithm import rs_image
import openpyxl
import glob


def saveExcel(Array, savePath):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(savePath)


def fill_zeros(arr):
    non_zero_indices = [i for i, num in enumerate(arr) if num != 0]

    if not non_zero_indices:
        # 如果数组中全为0，则不进行操作
        return arr

    for i in range(len(arr)):
        if arr[i] == 0:
            # 找到离当前位置最近的非零元素
            nearest_non_zero_index = min(non_zero_indices, key=lambda x: abs(x - i))
            arr[i] = arr[nearest_non_zero_index]

    return arr


conflict_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\conflict_data\\' \
                'group_region_conflict\\'
group_region_file = 'D:\\Drought_and_heat_wave_coupling\\data\\mask\\gw\\raster_filter\\'

group_mask_path_list = glob.glob(group_region_file + '*.tif')


array_list = []
for i in range(0, len(group_mask_path_list), 1):
    group_code = int(group_mask_path_list[i][-7:-4])
    print(group_code)

    conflict_path = conflict_file + str(int(group_code)) + '.xlsx'
    conflict = pd.read_excel(conflict_path)
    conflict_array = np.array(conflict)
    years = conflict_array[:, 2]
    side = conflict_array[:, 15]
    conflict_list = []
    for j in range(1989, 2021, 1):
        if np.sum(years == j) > 0:
            year_side_pos = np.where(years == j)[0]
            side_name = np.unique(side[year_side_pos])
            side_number = side_name.size
            conflict_list.append(side_number)
        else:
            conflict_list.append(0)

    conflict_list = fill_zeros(conflict_list)
    conflict_list = np.array([conflict_list]).T
    index = np.full((32, 1), int(group_code))
    event_array = np.append(index, conflict_list, axis=1)
    array_list.append(event_array)

array = array_list[0]
for i in range(1, len(array_list), 1):
    array = np.append(array, array_list[i], axis=0)

saveExcel(array, 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\exp_1989_2020\\'
                 'political_number.xlsx')
