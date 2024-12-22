import pandas as pd
import numpy as np
import openpyxl


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\CDHW_conflict.xlsx'
data = pd.read_excel(path)
data = np.array(data)
ad = data[:, 5]
index = np.where(ad < 0.05)
del_array = np.delete(data, index, axis=0)
saveExcel(del_array, 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\CDHW_conflict_ad5.xlsx')