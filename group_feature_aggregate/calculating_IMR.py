from pyrs.algorithm import supervision_classification as sc
from pyrs.algorithm import rs_image
from osgeo import gdal
import numpy as np
import glob
import openpyxl
import pandas as pd
from sklearn.linear_model import LinearRegression


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


group_region_file = 'D:\\Drought_and_heat_wave_coupling\\data\\mask\\imr_civilian\\raster_filter\\'
group_mask_path_list = glob.glob(group_region_file + '*.tif')
image_file = 'D:\\Drought_and_heat_wave_coupling\\data\\IMR_image\\'

image_list = glob.glob(image_file + '*.tif')

array_list = []
for i in range(0, len(group_mask_path_list), 1):
    group_code = int(group_mask_path_list[i][-7:-4])
    group_region_data = rs_image.Image(group_mask_path_list[i])
    mask_array = group_region_data.get_array(True, 1)
    print(group_code)

    value_list = []
    for j in range(0, len(image_list), 1):
        image = rs_image.Image(image_list[j])
        image_array = image.get_array(True, 1)
        nan_index = np.isnan(image_array)
        image_array[nan_index] = 0
        image_array = np.where(image_array < 0, 0, image_array)

        pos = np.where(mask_array == 1)
        group_mean = np.mean(image_array[pos])
        value_list.append(group_mean)

    year = np.arange(2001, 2018).reshape((-1, 1))
    value = np.array(value_list)
    regression_model = LinearRegression()
    regression_model.fit(year, value)
    year_future = np.arange(2018, 2021).reshape((-1, 1))
    future = regression_model.predict(year_future)

    area_value = np.append(value, future).reshape((-1, 1))

    index = np.full((20, 1), group_code)
    array = np.append(index, area_value, axis=1)
    array_list.append(array)
    print(array_list)

array = array_list[0]
for i in range(1, len(array_list), 1):
    array = np.append(array, array_list[i], axis=0)

saveExcel(array, 'D:\\Drought_and_heat_wave_coupling\\data\\IMR_image\\imr_civilian.xlsx')
