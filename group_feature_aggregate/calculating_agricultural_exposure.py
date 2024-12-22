from pyrs.algorithm import supervision_classification as sc
from pyrs.algorithm import rs_image
from osgeo import gdal
import numpy as np
import glob
from sklearn.linear_model import LinearRegression
import openpyxl
import pandas as pd


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


group_region_file = 'D:\\Drought_and_heat_wave_coupling\\data\\mask\\ac\\raster_filter\\'
group_mask_path_list = glob.glob(group_region_file + '*.tif')

image_file = 'D:\\Drought_and_heat_wave_coupling\\data\\landcover\\'
duration_file = 'D:\\Drought_and_heat_wave_coupling\\data\\CDHW_feature\\duration\\'

landcover_list = glob.glob(image_file + '*.tif')
duration_list = glob.glob(duration_file + '*.tif')

array_list = []
for i in range(0, len(group_mask_path_list), 1):
    group_code = int(group_mask_path_list[i][-7:-4])
    group_region_data = rs_image.Image(group_mask_path_list[i])
    mask_array = group_region_data.get_array(True, 1)
    print(group_code)

    rainfall_list = []
    irrigate_list = []
    for j in range(0, len(landcover_list), 1):
        image = rs_image.Image(landcover_list[j])
        image_array = image.get_array(True, 1)

        duration = rs_image.Image(duration_list[j])
        duration_array = duration.resample(4008, 8016)

        image_array = np.where(image_array == 11, 10, image_array)
        image_array = np.where(image_array == 12, 10, image_array)

        rainfall_area = np.sum((mask_array == 1) & (image_array == 10))
        rainfall_exposure_area = np.sum((mask_array == 1) & (image_array == 10) & (duration_array > 0))
        if rainfall_area == 0:
            rainfall_ratio = 0
        else:
            rainfall_ratio = rainfall_exposure_area / rainfall_area
        rainfall_list.append(rainfall_ratio)

        irrigate_area = np.sum((mask_array == 1) & (image_array == 20))
        irrigate_exposure_area = np.sum((mask_array == 1) & (image_array == 20) & (duration_array > 0))
        if irrigate_area == 0:
            irrigate_ratio = 0
        else:
            irrigate_ratio = irrigate_exposure_area / irrigate_area
        irrigate_list.append(irrigate_ratio)

    year = np.arange(2001, 2021).reshape((-1, 1))
    rainfall_value = np.array(rainfall_list).reshape((-1, 1))
    irrigate_value = np.array(irrigate_list).reshape((-1, 1))

    index = np.full((20, 1), group_code)
    array = np.append(index, rainfall_value, axis=1)
    array = np.append(array, irrigate_value, axis=1)
    array_list.append(array)

array = array_list[0]
for i in range(1, len(array_list), 1):
    array = np.append(array, array_list[i], axis=0)

saveExcel(array, 'D:\\Drought_and_heat_wave_coupling\\data\\cumNDVI\\exposure.xlsx')