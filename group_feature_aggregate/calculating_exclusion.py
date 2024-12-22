import pandas as pd
import numpy as np
import openpyxl
import glob


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


select_table_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\conflict_data\\' \
                    'select_tabel.xlsx'
exclusion_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\exclusion\\EPR-2021.xlsx'
save_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\exclusion\\group\\'

select_table = pd.read_excel(select_table_path)
select_index = np.array(select_table['group index'])
select_statename = np.array(select_table['statename'])
select_group = np.array(select_table['group'])

exclusion_table = pd.read_excel(exclusion_path)
exclusion_statename = np.array(exclusion_table['statename'])
exclusion_group = np.array(exclusion_table['group'])
exclusion_time = np.array(exclusion_table['to'])
exclusion_level = np.array(exclusion_table['status'])

group_region_file = 'D:\\Drought_and_heat_wave_coupling\\data\\mask\\gw\\raster_filter\\'

group_mask_path_list = glob.glob(group_region_file + '*.tif')

group_code_list = []

for i in range(0, len(group_mask_path_list), 1):
    group_code = int(group_mask_path_list[i][-7:-4])
    group_code_list.append(group_code)

group_code_array = np.array(group_code_list)

array_list = []
for i in range(0, select_index.shape[0], 1):
    group_statename = select_statename[i]
    group = select_group[i]
    index = select_index[i]
    select_group_pos = np.where((exclusion_statename == group_statename) & (exclusion_group == group))[0]

    if np.where(group_code_array == index)[0].size == 0:
        continue

    level_array = np.zeros((32, 1))
    start = 0
    for j in range(0, select_group_pos.shape[0], 1):
        time = exclusion_time[select_group_pos[j]]
        level = exclusion_level[select_group_pos[j]]
        if level == 'MONOPOLY':
            level = 1
        elif level == 'DOMINANT':
            level = 2
        elif level == 'SENIOR PARTNER':
            level = 3
        elif level == 'JUNIOR PARTNER':
            level = 4
        elif level == 'IRRELEVANT':
            level = 5
        elif level == 'SELF-EXCLUSION':
            level = 6
        elif level == 'POWERLESS':
            level = 7
        elif level == 'DISCRIMINATED':
            level = 8
        else:
            level = 9

        if time >= 1989:
            end = time - 1988
            if end < 32:
                level_array[start:end] = level
                start = end
            else:
                level_array[start:] = level
                break

    index_array = np.full((32, 1), index)
    array = np.append(index_array, level_array, axis=1)
    array_list.append(array)

    #     if time >= 2001:
    #         end = time - 2000
    #         if end < 20:
    #             level_array[start:end] = level
    #             start = end
    #         else:
    #             level_array[start:] = level
    #             break
    #
    # index_array = np.full((20, 1), index)
    # array = np.append(index_array, level_array, axis=1)
    # array_list.append(array)

array = array_list[0]
for i in range(1, len(array_list), 1):
    array = np.append(array, array_list[i], axis=0)

saveExcel(array, 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\exp_1989_2020\\'
                 'exclusion.xlsx')