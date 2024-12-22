import pandas as pd
import numpy as np
from pyrs.algorithm import rs_image
import openpyxl
import glob


def saveExcel(Array, savePath):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(savePath)


save_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\country_team.xlsx'
group_region_file = 'D:\\Drought_and_heat_wave_coupling\\data\\mask\\country\\raster_filter\\'
team_save_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\' \
                         'conflict_data\\country_team_num.xlsx'

team = pd.read_excel(team_save_file)
team = np.array(team)

group_mask_path_list = glob.glob(group_region_file + '*.tif')
array_list = []

for i in range(0, len(group_mask_path_list), 1):
    group_code = int(group_mask_path_list[i][-7:-4])
    print(group_code)

    array = np.full((20, 1), team[i, 1])
    array_list.append(array)

new_array = array_list[0]
for i in range(1, len(array_list), 1):
    new_array = np.append(new_array, array_list[i], axis=0)

saveExcel(new_array, save_file)