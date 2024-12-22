import numpy as np
import pandas as pd
import openpyxl


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


select_table_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\conflict_data\\' \
                    'select_tabel.xlsx'
data_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\CDHW2\\with_neo.xlsx'
save_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\CDHW2\\mechanism\\'

continent_name = 'normal'

data_save_path = save_file + 'normal.xlsx'

select_table = pd.read_excel(select_table_path)
data = np.array(pd.read_excel(data_path))

continent = np.array(select_table['type'])
state = np.array(select_table['group index'])
pos = np.where(continent == continent_name)
state_number = np.unique(state[pos])
# print(state_number)

data_state_number = data[:, 1]
indices = np.where(~np.isin(data_state_number, state_number))[0]
new_data = np.delete(data, indices, axis=0)
saveExcel(new_data, data_save_path)