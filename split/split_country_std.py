import pandas as pd
import numpy as np
import openpyxl


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


def split_table(data, feature_name, condition, save_path):
    feature_array = np.array(data[feature_name])
    state_array = np.array(data['state'])
    data = np.array(data)
    state_index = np.unique(state_array)
    state_std_list = []
    for i in range(0, state_index.shape[0], 1):
        state_pos = np.where(state_array == state_index[i])
        state_feature_std = np.std(feature_array[state_pos])
        state_std_list.append(state_feature_std)

    if condition == 'p1':
        threshold = np.percentile(state_std_list, 25)
        pos = np.where(state_std_list < threshold)[0]
        state_value = state_index[pos]
        indices = np.where(~np.isin(state_array, state_value))[0]
        new_data = np.delete(data, indices, axis=0)
        saveExcel(new_data, save_path)

    if condition == 'p2':
        threshold1 = np.percentile(state_std_list, 50)
        threshold2 = np.percentile(state_std_list, 25)
        pos = np.where((state_std_list < threshold1) & (state_std_list >= threshold2))[0]
        state_value = state_index[pos]
        indices = np.where(~np.isin(state_array, state_value))[0]
        new_data = np.delete(data, indices, axis=0)
        saveExcel(new_data, save_path)

    if condition == 'p3':
        threshold1 = np.percentile(state_std_list, 75)
        threshold2 = np.percentile(state_std_list, 50)
        pos = np.where((state_std_list < threshold1) & (state_std_list >= threshold2))[0]
        state_value = state_index[pos]
        indices = np.where(~np.isin(state_array, state_value))[0]
        new_data = np.delete(data, indices, axis=0)
        saveExcel(new_data, save_path)

    if condition == 'p4':
        threshold = np.percentile(state_std_list, 75)
        pos = np.where(state_std_list <= threshold)[0]
        state_value = state_index[pos]
        indices = np.where(~np.isin(state_array, state_value))[0]
        new_data = np.delete(data, indices, axis=0)
        saveExcel(new_data, save_path)


if __name__ == '__main__':
    data_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\CDHW2\\' \
                'vulnerability_index_plus.xlsx'

    data_pd = pd.read_excel(data_path)
    name = 'hsvi'
    threshold_condition = 'p1'

    save_name = threshold_condition + '_' + name + '_state_plus.xlsx'
    save_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\CDHW2\\mechanism\\' + \
                save_name

    split_table(data_pd, name, threshold_condition, save_file)