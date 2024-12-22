import pandas as pd
import numpy as np
from pyrs.algorithm import rs_image
import openpyxl
import glob


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


# spei_file = 'D:\\Drought_and_heat_wave_coupling\\data\\SPEI_data\\spei\\'
# sample = rs_image.Image('D:\\Drought_and_heat_wave_coupling\\data\\SPEI_data\\spei\\1989\\01.tiff')
# year_list = glob.glob(spei_file + '*')
#
# num = 0
# average_spei = np.zeros((len(year_list) * 12, 360, 720))
# for i in range(0, len(year_list), 1):
#     month_list = glob.glob(year_list[i] + '\\*.tiff')
#     for j in range(0, len(month_list), 1):
#         image = rs_image.Image(month_list[j])
#         image_array = image.get_array(True, 1, np.float32)
#         average_spei[num, :, :] = image_array
#         num += 1
#
# average_spei = np.median(average_spei, axis=0)
# average_spei = np.where(average_spei > 2, 1000, average_spei)
# sample.save('D:\\Drought_and_heat_wave_coupling\\data\\SPEI_data\\average_spei_median.tif', average_spei)

output_raster_file2 = 'D:\\Drought_and_heat_wave_coupling\\data\\mask\\PDSI\\raster_filter\\'
spei = 'D:\\Drought_and_heat_wave_coupling\\data\\PDSI\\PDSI_Global.tif'

value_list = []

spei_array = rs_image.Image(spei).get_array(True, 1)
nan_pos = np.isnan(spei_array)
spei_array[nan_pos] = 1000

path_list = glob.glob(output_raster_file2 + '*.tif')
for i in range(0, len(path_list), 1):
    image = rs_image.Image(path_list[i]).get_array(True, 1)
    pos = np.where((image == 1) & (spei_array != 1000))
    average_spei = np.mean(spei_array[pos]) * 0.01
    if average_spei < -1:
        value = 'dry'
    elif average_spei > 1:
        value = 'wet'
    else:
        value = 'normal'
    value_list.append(value)
    # value_list.append(average_spei)

# value_list = np.array(value_list).reshape((-1, 1)) * 0.01
value_list = np.array(value_list).reshape((-1, 1))
saveExcel(value_list,
          'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\conflict_data\\dry_wet2.xlsx')