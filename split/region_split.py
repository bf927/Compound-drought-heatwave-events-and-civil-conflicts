import pandas as pd
import numpy as np
import openpyxl


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


def split_table(data, feature_name, condition, save_path):
    feature_array = np.array(data[feature_name])
    group_array = np.array(data['group'])
    data = np.array(data)
    group_index = np.unique(group_array)
    region_mean_feature_list = []
    for i in range(0, group_array.shape[0], 20):
        group_feature = feature_array[i:i + 20]
        mean_feature = np.mean(group_feature)
        region_mean_feature_list.append(mean_feature)

    if condition == 'p1':
        threshold = np.percentile(region_mean_feature_list, 25)
        pos = np.where(region_mean_feature_list < threshold)[0]
        group_value = group_index[pos]
        indices = np.where(~np.isin(group_array, group_value))[0]
        new_data = np.delete(data, indices, axis=0)
        saveExcel(new_data, save_path)

    if condition == 'p2':
        threshold1 = np.percentile(region_mean_feature_list, 50)
        threshold2 = np.percentile(region_mean_feature_list, 25)
        pos = np.where((region_mean_feature_list < threshold1) & (region_mean_feature_list >= threshold2))[0]
        group_value = group_index[pos]
        indices = np.where(~np.isin(group_array, group_value))
        new_data = np.delete(data, indices, axis=0)
        saveExcel(new_data, save_path)

    if condition == 'p3':
        threshold1 = np.percentile(region_mean_feature_list, 75)
        threshold2 = np.percentile(region_mean_feature_list, 50)
        pos = np.where((region_mean_feature_list < threshold1) & (region_mean_feature_list >= threshold2))[0]
        group_value = group_index[pos]
        indices = np.where(~np.isin(group_array, group_value))
        new_data = np.delete(data, indices, axis=0)
        saveExcel(new_data, save_path)

    if condition == 'p4':
        threshold = np.percentile(region_mean_feature_list, 75)
        pos = np.where(region_mean_feature_list >= threshold)[0]
        group_value = group_index[pos]
        indices = np.where(~np.isin(group_array, group_value))[0]
        new_data = np.delete(data, indices, axis=0)
        saveExcel(new_data, save_path)

    if condition == 'p1_p3':
        threshold = np.percentile(region_mean_feature_list, 75)
        pos = np.where(region_mean_feature_list < threshold)[0]
        group_value = group_index[pos]
        indices = np.where(~np.isin(group_array, group_value))[0]
        new_data = np.delete(data, indices, axis=0)
        saveExcel(new_data, save_path)

    if condition == 'p3_p4':
        threshold = np.percentile(region_mean_feature_list, 50)
        pos = np.where(region_mean_feature_list >= threshold)[0]
        group_value = group_index[pos]
        indices = np.where(~np.isin(group_array, group_value))[0]
        new_data = np.delete(data, indices, axis=0)
        saveExcel(new_data, save_path)


if __name__ == '__main__':
    data_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\CDHW\\main_imr.xlsx'

    data_pd = pd.read_excel(data_path)
    name = 'hs_ssvi'
    threshold_condition = 'p4'

    save_name = threshold_condition + '_' + name + '.xlsx'
    save_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\CDHW\\' + save_name

    split_table(data_pd, name, threshold_condition, save_file)