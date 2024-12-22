import pandas as pd
import numpy as np
import openpyxl


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\CDHW\\main0.xlsx'
data = pd.read_excel(path)

hsvi = np.array(data['hs_encode'])
ssvi = np.array(data['ss_encode'])

union = np.zeros(hsvi.shape)

# for i in range(0, union.shape[0], 1):
#     if (hsvi[i] == 1) and (ssvi[i] == 1):
#         union[i] = 1
#     elif (hsvi[i] == 1) and (ssvi[i] == 2):
#         union[i] = 2
#     elif (hsvi[i] == 1) and (ssvi[i] == 3):
#         union[i] = 3
#     elif (hsvi[i] == 1) and (ssvi[i] == 4):
#         union[i] = 4
#     elif (hsvi[i] == 2) and (ssvi[i] == 1):
#         union[i] = 5
#     elif (hsvi[i] == 2) and (ssvi[i] == 2):
#         union[i] = 6
#     elif (hsvi[i] == 2) and (ssvi[i] == 3):
#         union[i] = 7
#     elif (hsvi[i] == 2) and (ssvi[i] == 4):
#         union[i] = 8
#     elif (hsvi[i] == 3) and (ssvi[i] == 1):
#         union[i] = 9
#     elif (hsvi[i] == 3) and (ssvi[i] == 2):
#         union[i] = 10
#     elif (hsvi[i] == 3) and (ssvi[i] == 3):
#         union[i] = 11
#     elif (hsvi[i] == 3) and (ssvi[i] == 4):
#         union[i] = 12
#     elif (hsvi[i] == 4) and (ssvi[i] == 1):
#         union[i] = 13
#     elif (hsvi[i] == 4) and (ssvi[i] == 2):
#         union[i] = 14
#     elif (hsvi[i] == 4) and (ssvi[i] == 3):
#         union[i] = 15
#     elif (hsvi[i] == 4) and (ssvi[i] == 4):
#         union[i] = 16

for i in range(0, union.shape[0], 1):
    if (hsvi[i] == 1) and (ssvi[i] == 1):
        union[i] = 1
    elif (hsvi[i] == 1) and (ssvi[i] == 2):
        union[i] = 2
    elif (hsvi[i] == 2) and (ssvi[i] == 1):
        union[i] = 3
    elif (hsvi[i] == 2) and (ssvi[i] == 2):
        union[i] = 4


saveExcel(union.reshape((-1, 1)), 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\CDHW\\'
                                  'union0.xlsx')