import pandas as pd
import numpy as np
import openpyxl


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


def encode_array5(array):
    p20 = np.percentile(array, 20)
    p40 = np.percentile(array, 40)
    p60 = np.percentile(array, 60)
    p80 = np.percentile(array, 80)

    print(p20)
    print(p40)
    print(p60)
    print(p80)

    pos0 = np.where(array <= p20)[0]
    pos1 = np.where((array > p20) & (array <= p40))[0]
    pos2 = np.where((array > p40) & (array <= p60))[0]
    pos3 = np.where((array > p60) & (array <= p80))[0]
    pos4 = np.where(array > p80)[0]

    new_array = np.zeros(array.shape[0])
    new_array[pos0] = 1
    new_array[pos1] = 2
    new_array[pos2] = 3
    new_array[pos3] = 4
    new_array[pos4] = 5
    return new_array


def encode_array4(array):
    p25 = np.percentile(array, 25)
    p50 = np.percentile(array, 50)
    p75 = np.percentile(array, 75)

    print(p25)
    print(p50)
    print(p75)

    pos0 = np.where(array <= p25)[0]
    pos1 = np.where((array > p25) & (array <= p50))[0]
    pos2 = np.where((array > p50) & (array <= p75))[0]
    pos3 = np.where(array > p75)[0]

    new_array = np.zeros(array.shape[0])
    new_array[pos0] = 1
    new_array[pos1] = 2
    new_array[pos2] = 3
    new_array[pos3] = 4
    return new_array


def encode_array2(array):
    p50 = np.percentile(array, 50)

    pos0 = np.where(array <= p50)[0]
    pos1 = np.where(array > p50)[0]

    new_array = np.zeros(array.shape[0])
    new_array[pos0] = 1
    new_array[pos1] = 2
    return new_array


def encode_array3(array):
    p33 = np.percentile(array, 33)
    p66 = np.percentile(array, 66)

    pos0 = np.where(array <= p33)[0]
    pos1 = np.where((array > p33) & (array <= p66))[0]
    pos2 = np.where(array > p66)[0]

    new_array = np.zeros(array.shape[0])
    new_array[pos0] = 1
    new_array[pos1] = 2
    new_array[pos2] = 3

    return new_array


path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\exp_1989_2020\\vi.xlsx'
data = pd.read_excel(path)
var = np.array(data['vi'])

new_var = encode_array4(var)

array2 = np.zeros((var.shape[0], 1))
array2[:, 0] = new_var

saveExcel(array2,
          'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\exp_1989_2020\\vi_encode.xlsx')

# path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\plt\\vi_mapping\\4p2.xlsx'
#
# data = pd.read_excel(path)
# hsvi = np.array(data['hsvi'])
# ssvi = np.array(data['ssvi'])
# # duration = np.array(data['duration'])
#
# print('hsvi')
# new_hsvi = encode_array(hsvi)
# print('ssvi')
# new_ssvi = encode_array(ssvi)
# print('duration')
# # new_duration = encode_array(duration)
#
# array2 = np.zeros((hsvi.shape[0], 3))
# array2[:, 0] = new_hsvi
# array2[:, 1] = new_ssvi
# # array2[:, 2] = new_duration
#
# # expanded_arrays = np.repeat(array2, 20, axis=0)
# # print(expanded_arrays.shape)
# # saveExcel(array2,
# #           'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\CDHW2\\encode.xlsx')
