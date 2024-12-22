import numpy as np
import pandas as pd
import openpyxl


def saveExcel(array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(array)):
        for jj in range(len(array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = array[ii][jj]  # 写入数据
    workbook.save(save_path)


total_event_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\' \
                   'data\\conflict_data\\nostate_conflict.xlsx'
ADC2EPR_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\' \
               'data\\conflict_data\\ACD2EPR-2021.xlsx'
group_region_save_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\' \
                         'conflict_data\\side_b_conflict\\'
select_tabel_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\' \
                   'data\\conflict_data\\select_tabel_side_b.xlsx'

total_event = pd.read_excel(total_event_path)

# 全部事件的政治团体
total_sideb = total_event['side_b']
total_sideb = np.array(total_sideb[1:])

total_event = np.array(total_event)
column_name = np.array([total_event[0, :]])
total_event = total_event[1:, :]

ADC2EPR = pd.read_excel(ADC2EPR_path)

# 国家
ADC2EPR_statename = ADC2EPR['statename']
ADC2EPR_statename = np.array(ADC2EPR_statename)
unique_statename = np.unique(ADC2EPR_statename)

# 种族
ADC2EPR_group = ADC2EPR['group']
ADC2EPR_group = np.array(ADC2EPR_group)

# 政治团体
ADC2EPR_sideb = ADC2EPR['sideb']
ADC2EPR_sideb = np.array(ADC2EPR_sideb)

state_list = []
group_list = []
number = 0
# 遍历每个国家中的各个种族
for i in range(0, unique_statename.shape[0], 1):
    print(unique_statename[i])
    state_element_pos = np.where(ADC2EPR_statename == unique_statename[i])[0]
    state_group = ADC2EPR_group[state_element_pos]  # 每个国家内的种族
    state_sideb = ADC2EPR_sideb[state_element_pos]  # 每个国家内的政治团体
    unique_state_group = np.unique(state_group)  # 每个国家内的种族名称
    # print(unique_state_group)

    for j in range(0, unique_state_group.shape[0], 1):
        state_group_pos = np.where(state_group == unique_state_group[j])[0]  # 获得每个国家每个种族的索引，可能大于1个
        print(unique_state_group[j])
        # print(state_group_pos.shape[0])

        group_conflict_info_list = []  # 存放同一种族不同政治团体的战争信息
        for k in range(0, state_group_pos.shape[0], 1):  # 同一种族可能对应多个政治团体
            sideb_name = state_sideb[state_group_pos[k]]
            print(sideb_name)
            sideb_name_pos = np.where(total_sideb == sideb_name)[0]
            if sideb_name_pos.size == 0:
                continue
            else:
                group_conflict_info = total_event[sideb_name_pos, :]
                if group_conflict_info.ndim == 1:
                    group_conflict_info = np.array([group_conflict_info])
                # print(group_conflict_info)
                group_conflict_info_list.append(group_conflict_info)

        if len(group_conflict_info_list) != 0:
            # print(group_conflict_info_list)
            group_conflict_info_array = group_conflict_info_list[0]
            for k in range(1, len(group_conflict_info_list), 1):
                group_conflict_info_array = np.append(group_conflict_info_array,
                                                      group_conflict_info_list[k],
                                                      axis=0)
            group_conflict_info_array = np.append(column_name, group_conflict_info_array, axis=0)
            group_region_save_path = group_region_save_file + str(number) + '.xlsx'
            # saveExcel(group_conflict_info_array, group_region_save_path)
            number += 1

            state_list.append(unique_statename[i])
            group_list.append(unique_state_group[j])

state_list = np.array([state_list])
group_list = np.array([group_list])
select_tabel = np.append(state_list, group_list, axis=0).T
# saveExcel(select_tabel, select_tabel_path)