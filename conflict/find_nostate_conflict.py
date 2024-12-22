import numpy as np
import pandas as pd
import openpyxl


def saveExcel(array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(array)):
        for jj in range(len(array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = array[ii][jj]  # 写入数据
    workbook.save(save_path)


def find_equal_row(a, b):
    # 获取数组a和数组b的第一列
    first_col_a = a[:, 0]
    first_col_b = b[:, 0]

    # 找到a和b中完全相同的元素
    matching_values = np.intersect1d(first_col_a, first_col_b)

    # 初始化一个空列表来存储匹配元素在a中的索引
    matching_indices = []

    # 遍历匹配的元素
    for value in matching_values:
        # 获取匹配元素在a中的索引，并添加到列表中
        indices = np.where(first_col_a == value)[0]
        matching_indices.extend(indices)

    # 使用索引获取匹配的行
    matching_rows = a[matching_indices, :]

    return matching_rows


side_a_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\' \
              'data\\conflict_data\\select_tabel_side_a.xlsx'
side_b_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\' \
              'data\\conflict_data\\select_tabel_side_b.xlsx'
side_a_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\' \
                         'conflict_data\\side_a_conflict\\'
side_b_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\' \
                         'conflict_data\\side_b_conflict\\'
save_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\' \
                         'conflict_data\\nostate_conflict\\'
total_event_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\' \
                   'data\\conflict_data\\nostate_conflict.xlsx'

total_event = pd.read_excel(total_event_path)

total_event = np.array(total_event)
column_name = np.array([total_event[0, :]])

side_a = pd.read_excel(side_a_path)
side_b = pd.read_excel(side_b_path)

side_a_id = np.array(side_a['group index'])
side_b_id = np.array(side_b['group index'])

side_a_main_id = np.array(side_a['main index'])
side_b_main_id = np.array(side_b['main index'])

side_a_country = np.array(side_a['statename'])
side_b_country = np.array(side_b['statename'])


for i in range(0, side_a_id.shape[0], 1):
    print('i:' + str(i))
    for j in range(0, side_b_id.shape[0], 1):
        print(j)
        a_group = np.array(pd.read_excel(side_a_file + str(i) + '.xlsx'))
        b_group = np.array(pd.read_excel(side_b_file + str(j) + '.xlsx'))
        result_array = find_equal_row(a_group, b_group)

        if result_array.size != 0:
            a_formatted = "{:03}".format(int(side_a_main_id[i]))
            b_formatted = "{:03}".format(int(side_b_main_id[j]))
            save_path = save_file + a_formatted + '_' + b_formatted + '.xlsx'
            if result_array.size == 1:
                result_array = np.array([result_array])
            result_array = np.append(column_name, result_array, axis=0)
            saveExcel(result_array, save_path)