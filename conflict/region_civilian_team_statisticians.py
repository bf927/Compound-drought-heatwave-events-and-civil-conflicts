import numpy as np
import pandas as pd
import openpyxl


def saveExcel(array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(array)):
        for jj in range(len(array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = array[ii][jj]  # 写入数据
    workbook.save(save_path)


ADC2EPR_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\' \
               'data\\conflict_data\\ACD2EPR-2021.xlsx'
team_save_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\' \
                         'conflict_data\\country_team_num.xlsx'
select_tabel_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\' \
                    'data\\conflict_data\\select_tabel_civilian.xlsx'

select_tabel = pd.read_excel(select_tabel_path)
ADC2EPR = pd.read_excel(ADC2EPR_path)

ADC2EPR_statename = ADC2EPR['statename']
ADC2EPR_statename = np.array(ADC2EPR_statename)


select_tabel_country = np.array(select_tabel['country'])
team_list = []
for i in range(0, select_tabel_country.shape[0], 1):
    pos = np.where(ADC2EPR_statename == select_tabel_country[i])[0]
    team_list.append(pos.size)

team_list = np.array(team_list).reshape((-1, 1))
saveExcel(team_list, team_save_file)
