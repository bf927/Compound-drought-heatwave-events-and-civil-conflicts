import numpy as np
import pandas as pd
import openpyxl


def saveExcel(array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(array)):
        for jj in range(len(array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = array[ii][jj]  # 写入数据
    workbook.save(save_path)


total_event_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\' \
                   'data\\conflict_data\\civilian_Conflict.xlsx'
group_region_save_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\' \
                         'conflict_data\\country_region_conflict\\'
select_tabel_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\' \
                   'data\\conflict_data\\select_tabel_civilian.xlsx'

total_event = pd.read_excel(total_event_path)
select_tabel = pd.read_excel(select_tabel_path)

# 全部事件的政治团体
total_sideb = total_event['country']
total_sideb = np.array(total_sideb[1:])

total_event = np.array(total_event)
column_name = np.array([total_event[0, :]])
total_event = total_event[1:, :]

select_tabel_country = np.array(select_tabel['country'])
for i in range(0, select_tabel_country.shape[0], 1):
    pos = np.where(total_sideb == select_tabel_country[i])[0]
    country_conflict = np.append(column_name, total_event[pos, :], axis=0)
    save_path = group_region_save_file + str(i) + '.xlsx'
    saveExcel(country_conflict, save_path)