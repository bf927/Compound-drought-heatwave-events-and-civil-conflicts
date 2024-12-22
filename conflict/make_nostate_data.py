import numpy as np
import pandas as pd
import openpyxl
import glob


def saveExcel(array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(array)):
        for jj in range(len(array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = array[ii][jj]  # 写入数据
    workbook.save(save_path)


data_file = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\' \
            'conflict_data\\nostate_conflict\\'
group_conflict_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\CDHW2\\' \
                      'vulnerability_index.xlsx'
conflict_data = pd.read_excel(group_conflict_path)
group_index = np.array(conflict_data['group'])
conflict_data = np.array(conflict_data)[:, 0:-4]

data_list = glob.glob(data_file + '*.xlsx')
array_list = []

num = 169

for i in range(0, len(data_list), 1):
    index = data_list[i][-12:-5]
    side_a_id = int(index[0:3])
    side_b_id = int(index[4:])
    side_a_pos = np.where(group_index == side_a_id)[0]
    side_b_pos = np.where(group_index == side_b_id)[0]

    if (side_a_pos.size != 0) and (side_b_pos.size != 0):
        bi_side_data = pd.read_excel(data_list[i])
        years = bi_side_data['year']
        conflict_list = []
        for j in range(2001, 2021, 1):
            if np.sum(years == j) > 0:
                conflict_list.append(np.sum(years == j))
            else:
                conflict_list.append(0)

        conflict_list = np.array([conflict_list]).T
        side_a_data = conflict_data[side_a_pos, :]
        side_b_data = conflict_data[side_b_pos, :]
        new_data = (side_a_data + side_b_data) / 2
        new_data = np.append(new_data, conflict_list, axis=1)
        new_data[:, 1] = np.full((20,), num)
        array_list.append(new_data)
        num += 1
        # print(new_data.shape)
    else:
        continue

new_array = array_list[0]
for i in range(1, len(array_list), 1):
    new_array = np.append(new_array, array_list[i], axis=0)

saveExcel(new_array, 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\'
          'CDHW_nostate\\CDHW_nostate2.xlsx')
