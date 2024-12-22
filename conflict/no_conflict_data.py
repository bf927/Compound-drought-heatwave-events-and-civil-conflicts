import numpy as np
import pandas as pd
import openpyxl


def saveExcel(Array, save_path):
    workbook = openpyxl.Workbook()
    sheet0 = workbook.create_sheet(index=0)  # 创建sheet0

    for ii in range(len(Array)):
        for jj in range(len(Array[ii])):
            sheet0.cell(ii + 1, jj + 1).value = Array[ii][jj]  # 写入数据
    workbook.save(save_path)


data_path = 'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\death_HW\\' \
            'main_hw_death4.xlsx'
data_pd = pd.read_excel(data_path)

data = np.array(data_pd)
array_list = []
for i in range(0, data.shape[0], 20):
    group_array = data[i:i+20, :]
    conflict = group_array[:, -1]
    if np.sum(conflict) != 0:
        array_list.append(group_array)

print(len(array_list))
new_array = array_list[0]
for i in range(1, len(array_list), 1):
    new_array = np.append(new_array, array_list[i], axis=0)
    # new_array = new_array + array_list[i]

# new_array = new_array / len(array_list)
saveExcel(new_array,
          'C:\\Users\\Administrator\\Desktop\\Drought_and_heat_wave_coupling\\data\\death_HW\\'
          'have_conflict.xlsx')